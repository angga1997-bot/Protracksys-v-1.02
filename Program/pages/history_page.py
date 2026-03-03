"""
pages/history_page.py - History Data Page (Read from JSON)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date, timedelta
from PIL import Image, ImageTk
import os
import json

from pages.base_page import BasePage
from config import COLORS, CAPTURED_DIR


class HistoryPage(BasePage):
    """Halaman untuk melihat dan export history data"""
    
    def __init__(self, parent, controller):
        super().__init__(parent, controller)
        self.photo_refs = {}
        self.current_page = 0
        self.page_size = 50
        self.total_records = 0
        self.all_data = []
        self.filtered_data = []
        self._create_widgets()
    
    def _create_widgets(self):
        """Membuat widget"""
        # Header
        self.create_header("📜 History Data", "Search and export production data")
        
        # Search section
        self._create_search_section()
        
        # Filter section
        self._create_filter_section()
        
        # Stats section
        self._create_stats_section()
        
        # Data section
        self._create_data_section()
        
        # Pagination
        self._create_pagination()
    
    def _create_search_section(self):
        """Text search section"""
        search_frame = tk.Frame(self, bg=self.colors["bg_card"])
        search_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        inner = tk.Frame(search_frame, bg=self.colors["bg_card"])
        inner.pack(fill="x", padx=20, pady=15)
        
        # Search label
        tk.Label(
            inner,
            text="🔍 Search in Table:",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_primary"]
        ).pack(side="left")
        
        # Search entry
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(
            inner,
            textvariable=self.search_var,
            font=("Segoe UI", 11),
            bg=self.colors["bg_hover"],
            fg=self.colors["text_primary"],
            insertbackground=self.colors["text_primary"],
            relief="flat",
            width=40
        )
        self.search_entry.pack(side="left", padx=10, ipady=6)
        self.search_entry.bind("<Return>", lambda e: self._apply_search())
        self.search_entry.bind("<KeyRelease>", self._on_search_key_release)
        
        # Search options
        options_frame = tk.Frame(inner, bg=self.colors["bg_card"])
        options_frame.pack(side="left", padx=10)
        
        # Case sensitive
        self.case_sensitive_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            options_frame,
            text="Aa",
            variable=self.case_sensitive_var,
            font=("Segoe UI", 9),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"],
            activebackground=self.colors["bg_card"],
            selectcolor=self.colors["bg_hover"],
            command=self._apply_search
        ).pack(side="left", padx=2)
        
        # Search column
        tk.Label(
            options_frame,
            text="Column:",
            font=("Segoe UI", 9),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"]
        ).pack(side="left", padx=(10, 2))
        
        self.search_column_var = tk.StringVar(value="All Columns")
        self.search_column_combo = ttk.Combobox(
            options_frame,
            textvariable=self.search_column_var,
            values=["All Columns"],
            width=15,
            state="readonly"
        )
        self.search_column_combo.pack(side="left", padx=2)
        self.search_column_combo.bind("<<ComboboxSelected>>", lambda e: self._apply_search())
        
        # Search button
        tk.Button(
            inner,
            text="🔍 Search",
            font=("Segoe UI", 10),
            bg=self.colors["accent"],
            fg="#1e1e2e",
            relief="flat",
            cursor="hand2",
            padx=15, pady=5,
            command=self._apply_search
        ).pack(side="left", padx=5)
        
        # Clear button
        tk.Button(
            inner,
            text="✖ Clear",
            font=("Segoe UI", 10),
            bg=self.colors["accent_red"],
            fg="#1e1e2e",
            relief="flat",
            cursor="hand2",
            padx=10, pady=5,
            command=self._clear_search
        ).pack(side="left", padx=5)
        
        # Search result
        self.search_result_label = tk.Label(
            inner,
            text="",
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["accent_green"]
        )
        self.search_result_label.pack(side="right", padx=10)
    
    def _create_filter_section(self):
        """Date filter section"""
        filter_frame = tk.Frame(self, bg=self.colors["bg_card"])
        filter_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        inner = tk.Frame(filter_frame, bg=self.colors["bg_card"])
        inner.pack(fill="x", padx=20, pady=15)
        
        # Date filters
        tk.Label(
            inner, text="📅 From:",
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"]
        ).pack(side="left")
        
        self.start_date_var = tk.StringVar(value=(date.today() - timedelta(days=30)).isoformat())
        start_entry = tk.Entry(
            inner,
            textvariable=self.start_date_var,
            font=("Segoe UI", 10),
            bg=self.colors["bg_hover"],
            fg=self.colors["text_primary"],
            insertbackground=self.colors["text_primary"],
            relief="flat", width=12
        )
        start_entry.pack(side="left", padx=5, ipady=4)
        
        tk.Label(
            inner, text="To:",
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"]
        ).pack(side="left", padx=(15, 0))
        
        self.end_date_var = tk.StringVar(value=date.today().isoformat())
        end_entry = tk.Entry(
            inner,
            textvariable=self.end_date_var,
            font=("Segoe UI", 10),
            bg=self.colors["bg_hover"],
            fg=self.colors["text_primary"],
            insertbackground=self.colors["text_primary"],
            relief="flat", width=12
        )
        end_entry.pack(side="left", padx=5, ipady=4)
        
        # Quick date buttons
        quick_frame = tk.Frame(inner, bg=self.colors["bg_card"])
        quick_frame.pack(side="left", padx=10)
        
        quick_dates = [
            ("Today", 0),
            ("7 Days", 7),
            ("30 Days", 30),
            ("All", -1)
        ]
        
        for text, days in quick_dates:
            tk.Button(
                quick_frame,
                text=text,
                font=("Segoe UI", 8),
                bg=self.colors["bg_hover"],
                fg=self.colors["text_primary"],
                relief="flat",
                cursor="hand2",
                padx=5, pady=2,
                command=lambda d=days: self._set_quick_date(d)
            ).pack(side="left", padx=2)
        
        # Load button
        tk.Button(
            inner, text="📥 Load Data",
            font=("Segoe UI", 10),
            bg=self.colors["accent"],
            fg="#1e1e2e", relief="flat",
            cursor="hand2", padx=15, pady=5,
            command=self._load_all_data
        ).pack(side="left", padx=15)
        
        # Export buttons
        tk.Button(
            inner, text="📥 Export CSV",
            font=("Segoe UI", 10),
            bg=self.colors["accent_green"],
            fg="#1e1e2e", relief="flat",
            cursor="hand2", padx=15, pady=5,
            command=self._export_csv
        ).pack(side="right", padx=5)
        
        tk.Button(
            inner, text="📥 Export Excel",
            font=("Segoe UI", 10),
            bg=self.colors["accent_orange"],
            fg="#1e1e2e", relief="flat",
            cursor="hand2", padx=15, pady=5,
            command=self._export_excel
        ).pack(side="right", padx=5)
    
    def _create_stats_section(self):
        """Section statistik"""
        stats_frame = tk.Frame(self, bg=self.colors["bg_card"])
        stats_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        inner = tk.Frame(stats_frame, bg=self.colors["bg_card"])
        inner.pack(fill="x", padx=20, pady=10)
        
        self.stats_label = tk.Label(
            inner,
            text="📊 Total: 0 records",
            font=("Segoe UI", 11),
            bg=self.colors["bg_card"],
            fg=self.colors["text_primary"]
        )
        self.stats_label.pack(side="left")
        
        self.filter_stats_label = tk.Label(
            inner,
            text="",
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["accent_yellow"]
        )
        self.filter_stats_label.pack(side="left", padx=20)
        
        self.date_range_label = tk.Label(
            inner,
            text="",
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"]
        )
        self.date_range_label.pack(side="right")
    
    def _create_data_section(self):
        """Section data"""
        data_frame = tk.Frame(self, bg=self.colors["bg_card"])
        data_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        self.tree_frame = tk.Frame(data_frame, bg=self.colors["bg_card"])
        self.tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(self.tree_frame, orient="vertical")
        x_scroll = ttk.Scrollbar(self.tree_frame, orient="horizontal")
        
        # Treeview
        self.tree = ttk.Treeview(
            self.tree_frame,
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )
        
        y_scroll.config(command=self.tree.yview)
        x_scroll.config(command=self.tree.xview)
        
        y_scroll.pack(side="right", fill="y")
        x_scroll.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        
        # Style
        style = ttk.Style()
        style.configure("Treeview", rowheight=30, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        style.map("Treeview", background=[("selected", self.colors["accent"])])
        
        # Context menu
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="📋 Copy Value", command=self._copy_selected_value)
        self.context_menu.add_command(label="📋 Copy Row", command=self._copy_selected_row)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="🔍 Search This", command=self._search_selected_value)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="✏️ Edit", command=self._edit_selected)
        self.context_menu.add_command(label="🗑️ Delete", command=self._delete_selected)
        
        self.tree.bind("<Button-3>", self._show_context_menu)
        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Control-c>", lambda e: self._copy_selected_value())
    
    def _create_pagination(self):
        """Pagination controls"""
        pag_frame = tk.Frame(self, bg=self.colors["bg_card"])
        pag_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        inner = tk.Frame(pag_frame, bg=self.colors["bg_card"])
        inner.pack(pady=10)
        
        self.first_btn = tk.Button(
            inner, text="⏮",
            font=("Segoe UI", 10),
            bg=self.colors["bg_hover"],
            fg=self.colors["text_primary"],
            relief="flat", cursor="hand2", width=3,
            command=self._first_page
        )
        self.first_btn.pack(side="left", padx=2)
        
        self.prev_btn = tk.Button(
            inner, text="◀",
            font=("Segoe UI", 10),
            bg=self.colors["bg_hover"],
            fg=self.colors["text_primary"],
            relief="flat", cursor="hand2", width=3,
            command=self._prev_page
        )
        self.prev_btn.pack(side="left", padx=2)
        
        self.page_label = tk.Label(
            inner, text="Page 1 of 1",
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["text_primary"],
            width=15
        )
        self.page_label.pack(side="left", padx=10)
        
        self.next_btn = tk.Button(
            inner, text="▶",
            font=("Segoe UI", 10),
            bg=self.colors["bg_hover"],
            fg=self.colors["text_primary"],
            relief="flat", cursor="hand2", width=3,
            command=self._next_page
        )
        self.next_btn.pack(side="left", padx=2)
        
        self.last_btn = tk.Button(
            inner, text="⏭",
            font=("Segoe UI", 10),
            bg=self.colors["bg_hover"],
            fg=self.colors["text_primary"],
            relief="flat", cursor="hand2", width=3,
            command=self._last_page
        )
        self.last_btn.pack(side="left", padx=2)
        
        # Page size
        tk.Label(
            inner, text="Per page:",
            font=("Segoe UI", 9),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"]
        ).pack(side="left", padx=(20, 5))
        
        self.page_size_var = tk.StringVar(value="50")
        page_size_combo = ttk.Combobox(
            inner,
            textvariable=self.page_size_var,
            values=["25", "50", "100", "200"],
            width=5,
            state="readonly"
        )
        page_size_combo.pack(side="left")
        page_size_combo.bind("<<ComboboxSelected>>", lambda e: self._on_page_size_change())
    
    # ================================================================
    # DATA LOADING - FROM JSON
    # ================================================================
    
    def _load_all_data(self):
        """Load data from table_data.json and SQLite Database"""
        # 1. Get current shift data from JSON (same source as Dashboard)
        table_data = self.controller.table_data
        columns = table_data.get("columns", [])
        current_rows = table_data.get("rows", [])
        
        # Get column names
        self.column_names = []
        for col in columns:
            if isinstance(col, dict):
                self.column_names.append(col.get("name", "Column"))
            else:
                self.column_names.append(col)
        
        # 2. Get historical data from Database
        # Filter dates for database query to optimize
        start_date = self.start_date_var.get()
        end_date = self.end_date_var.get()
        db_history = self.controller.db.get_production_history(
            start_date=start_date if start_date else None,
            end_date=end_date if end_date else None,
            limit=5000 # reasonable limit
        )
        
        # 3. Combine data
        self.all_data = []
        
        # Add current shift data
        for row_idx, row in enumerate(current_rows):
            record = {"_row_index": f"json_{row_idx}", "_source": "current"}
            for col_idx, col_name in enumerate(self.column_names):
                if col_idx < len(row):
                    record[col_name] = row[col_idx]
                else:
                    record[col_name] = ""
            self.all_data.append(record)
            
        # Add historical data from DB
        for db_row in db_history:
            record = {"_row_index": f"db_{db_row['id']}", "_source": "history"}
            # The 'data' field in db_row is a dict containing column values
            data_dict = db_row.get('data', {})
            for col_name in self.column_names:
                record[col_name] = data_dict.get(col_name, "")
            self.all_data.append(record)
        
        # Sort by date if possible (optional, but good for display)
        # self.all_data.sort(key=lambda x: str(x.get(self._get_date_column() or "", "")), reverse=True)
        
        # Apply date filter (already applied in DB query for DB data, but re-run for JSON)
        self._apply_date_filter()
        
        # Update search column options
        self._update_search_column_options()
        
        # Apply search if exists
        if self.search_var.get().strip():
            self._apply_search()
        else:
            self.filtered_data = self.all_data.copy()
            self.search_result_label.configure(text="")
            self.current_page = 0
            self._display_current_page()
    
    def _apply_date_filter(self):
        """Filter data by date range"""
        start_date = self.start_date_var.get()
        end_date = self.end_date_var.get()
        
        # Find date column
        date_col = None
        for col_name in self.column_names:
            if any(keyword in col_name.lower() for keyword in ["tanggal", "date", "tgl"]):
                date_col = col_name
                break
        
        if not date_col or start_date == "" or end_date == "":
            # No date filtering
            self.date_range_label.configure(text=f"📊 All data")
            return
        
        # Filter by date
        filtered = []
        for record in self.all_data:
            record_date = record.get(date_col, "")
            if record_date:
                # Extract date part (YYYY-MM-DD)
                record_date_str = str(record_date)[:10]
                if start_date <= record_date_str <= end_date:
                    filtered.append(record)
                elif start_date == "" or end_date == "":
                    filtered.append(record)
            else:
                # Include records without date
                filtered.append(record)
        
        self.all_data = filtered
        self.date_range_label.configure(text=f"📅 {start_date} - {end_date}")
    
    def _set_quick_date(self, days):
        """Set quick date range"""
        if days == -1:
            # All data
            self.start_date_var.set("")
            self.end_date_var.set("")
        else:
            self.end_date_var.set(date.today().isoformat())
            self.start_date_var.set((date.today() - timedelta(days=days)).isoformat())
        
        self._load_all_data()
    
    def _update_search_column_options(self):
        """Update search column options"""
        options = ["All Columns"] + self.column_names
        self.search_column_combo["values"] = options
        if self.search_column_var.get() not in options:
            self.search_column_var.set("All Columns")
    
    # ================================================================
    # SEARCH FUNCTIONS
    # ================================================================
    
    def _on_search_key_release(self, event):
        """Live search with delay"""
        if hasattr(self, '_search_after_id'):
            self.after_cancel(self._search_after_id)
        self._search_after_id = self.after(300, self._apply_search)
    
    def _apply_search(self):
        """Apply search filter"""
        search_text = self.search_var.get().strip()
        
        if not search_text:
            self.filtered_data = self.all_data.copy()
            self.search_result_label.configure(text="")
        else:
            case_sensitive = self.case_sensitive_var.get()
            search_column = self.search_column_var.get()
            
            self.filtered_data = []
            
            if not case_sensitive:
                search_text = search_text.lower()
            
            for record in self.all_data:
                if self._record_matches(record, search_text, case_sensitive, search_column):
                    self.filtered_data.append(record)
            
            found_count = len(self.filtered_data)
            total_count = len(self.all_data)
            
            if found_count == 0:
                self.search_result_label.configure(
                    text="❌ Not found",
                    fg=self.colors["accent_red"]
                )
            else:
                self.search_result_label.configure(
                    text=f"✅ {found_count} of {total_count}",
                    fg=self.colors["accent_green"]
                )
        
        self.current_page = 0
        self._display_current_page()
    
    def _record_matches(self, record, search_text, case_sensitive, search_column):
        """Check if record matches search"""
        if search_column == "All Columns":
            columns_to_search = self.column_names
        else:
            columns_to_search = [search_column]
        
        for col in columns_to_search:
            value = str(record.get(col, ""))
            if not case_sensitive:
                value = value.lower()
            
            if search_text in value:
                return True
        
        return False
    
    def _clear_search(self):
        """Clear search"""
        self.search_var.set("")
        self.search_result_label.configure(text="")
        self.filtered_data = self.all_data.copy()
        self.current_page = 0
        self._display_current_page()
    
    # ================================================================
    # DISPLAY FUNCTIONS
    # ================================================================
    
    def _display_current_page(self):
        """Display current page"""
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Pagination
        self.page_size = int(self.page_size_var.get())
        self.total_records = len(self.filtered_data)
        total_pages = max(1, (self.total_records + self.page_size - 1) // self.page_size)
        
        if self.current_page >= total_pages:
            self.current_page = max(0, total_pages - 1)
        
        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, self.total_records)
        page_data = self.filtered_data[start_idx:end_idx]
        
        if not page_data and not self.column_names:
            self.stats_label.configure(text="📊 No data")
            self._update_pagination_buttons(0, 1)
            return
        
        # Configure columns
        display_columns = ["#"] + self.column_names
        self.tree["columns"] = display_columns
        self.tree["show"] = "headings"
        
        for col in display_columns:
            self.tree.heading(col, text=col, command=lambda c=col: self._sort_by_column(c))
            width = 60 if col == "#" else 120
            self.tree.column(col, width=width, anchor="center")
        
        # Insert data
        search_text = self.search_var.get().strip().lower() if self.search_var.get().strip() else None
        
        for idx, record in enumerate(page_data):
            row_num = start_idx + idx + 1
            values = [row_num] + [record.get(col, "") for col in self.column_names]
            
            # Highlight matching
            tags = ()
            if search_text:
                for v in values:
                    if search_text in str(v).lower():
                        tags = ("match",)
                        break
            
            self.tree.insert("", "end", values=values, tags=tags, iid=str(record.get("_row_index", idx)))
        
        # Highlight style
        self.tree.tag_configure("match", background=self.colors["accent_yellow"], foreground="#1e1e2e")
        
        # Update stats
        if len(self.filtered_data) != len(self.all_data):
            self.filter_stats_label.configure(
                text=f"(Filtered: {len(self.filtered_data)} dari {len(self.all_data)})"
            )
        else:
            self.filter_stats_label.configure(text="")
        
        self.stats_label.configure(
            text=f"📊 {start_idx + 1}-{end_idx} of {self.total_records} records"
        )
        
        self._update_pagination_buttons(self.current_page, total_pages)
    
    def _update_pagination_buttons(self, current_page, total_pages):
        """Update pagination buttons"""
        self.page_label.configure(text=f"Page {current_page + 1} / {total_pages}")
        
        can_prev = current_page > 0
        can_next = current_page < total_pages - 1
        
        self.first_btn.configure(state="normal" if can_prev else "disabled")
        self.prev_btn.configure(state="normal" if can_prev else "disabled")
        self.next_btn.configure(state="normal" if can_next else "disabled")
        self.last_btn.configure(state="normal" if can_next else "disabled")
    
    def _sort_by_column(self, column):
        """Sort by column"""
        if column == "#":
            return
        
        if not hasattr(self, '_sort_reverse'):
            self._sort_reverse = False
        self._sort_reverse = not self._sort_reverse
        
        def get_key(record):
            val = record.get(column, "")
            # Try numeric sort
            try:
                return (0, float(str(val).replace(",", "").replace(".", "")))
            except:
                return (1, str(val).lower())
        
        self.filtered_data.sort(key=get_key, reverse=self._sort_reverse)
        self._display_current_page()
    
    # ================================================================
    # PAGINATION
    # ================================================================
    
    def _first_page(self):
        self.current_page = 0
        self._display_current_page()
    
    def _prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self._display_current_page()
    
    def _next_page(self):
        total_pages = max(1, (self.total_records + self.page_size - 1) // self.page_size)
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self._display_current_page()
    
    def _last_page(self):
        total_pages = max(1, (self.total_records + self.page_size - 1) // self.page_size)
        self.current_page = total_pages - 1
        self._display_current_page()
    
    def _on_page_size_change(self):
        self.page_size = int(self.page_size_var.get())
        self.current_page = 0
        self._display_current_page()
    
    # ================================================================
    # CONTEXT MENU ACTIONS
    # ================================================================
    
    def _show_context_menu(self, event):
        """Show context menu"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self._selected_column = self.tree.identify_column(event.x)
            self.context_menu.post(event.x_root, event.y_root)
    
    def _on_double_click(self, event):
        """Double click to view image or search"""
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if item and column:
            col_idx = int(column.replace("#", "")) - 1
            values = self.tree.item(item, 'values')
            
            if 0 < col_idx < len(values):
                value = str(values[col_idx])
                if not value:
                    return
                
                # Cek apakah kolom ini tipe image
                columns_config = self.controller.table_data.get("columns", [])
                cfg_idx = col_idx - 1
                if 0 <= cfg_idx < len(columns_config):
                    col_cfg = columns_config[cfg_idx]
                    if isinstance(col_cfg, dict) and col_cfg.get("data_source") == "image":
                        folder = col_cfg.get("image_config", {}).get("folder_path", "").strip()
                        if not folder:
                            t_idx = col_cfg.get("image_config", {}).get("trigger_index", 0)
                            trigs = self.controller.data_manager.trigger_config.get("image_triggers", [])
                            if 0 <= t_idx - 1 < len(trigs):
                                folder = trigs[t_idx - 1].get("folder_path", "")
                            elif trigs:
                                folder = trigs[0].get("folder_path", "")
                        
                        import os
                        img_path = os.path.join(folder, value) if folder else ""
                        if img_path and os.path.isfile(img_path):
                            if hasattr(os, 'startfile'):
                                os.startfile(img_path)
                            return
                
                # Fallback: search
                self.search_var.set(value)
                self._apply_search()
    
    def _copy_selected_value(self):
        """Copy selected cell"""
        selection = self.tree.selection()
        if not selection or not hasattr(self, '_selected_column'):
            return
        
        col_idx = int(self._selected_column.replace("#", "")) - 1
        values = self.tree.item(selection[0], 'values')
        
        if 0 <= col_idx < len(values):
            value = str(values[col_idx])
            self.clipboard_clear()
            self.clipboard_append(value)
            self.search_result_label.configure(text="📋 Copied!", fg=self.colors["accent"])
    
    def _copy_selected_row(self):
        """Copy entire row"""
        selection = self.tree.selection()
        if not selection:
            return
        
        values = self.tree.item(selection[0], 'values')
        row_text = "\t".join([str(v) for v in values])
        
        self.clipboard_clear()
        self.clipboard_append(row_text)
        self.search_result_label.configure(text="📋 Row copied!", fg=self.colors["accent"])
    
    def _search_selected_value(self):
        """Search selected cell value"""
        selection = self.tree.selection()
        if not selection or not hasattr(self, '_selected_column'):
            return
        
        col_idx = int(self._selected_column.replace("#", "")) - 1
        values = self.tree.item(selection[0], 'values')
        
        if 0 <= col_idx < len(values):
            value = str(values[col_idx])
            if value:
                self.search_var.set(value)
                self._apply_search()
    
    def _edit_selected(self):
        """Edit selected row"""
        selection = self.tree.selection()
        if not selection:
            return
        
        row_index = int(selection[0])
        values = self.tree.item(selection[0], 'values')
        
        # Open edit dialog
        dialog = EditRowDialog(self, self.column_names, list(values[1:]), self.colors)
        self.wait_window(dialog)
        
        if dialog.result:
            # Update data
            rows = self.controller.table_data.get("rows", [])
            if row_index < len(rows):
                rows[row_index] = dialog.result
                self.controller.save_table_data()
                self._load_all_data()
                messagebox.showinfo("Success", "✅ Data updated successfully!")
    
    def _delete_selected(self):
        """Delete selected row"""
        selection = self.tree.selection()
        if not selection:
            return
        
        count = len(selection)
        if messagebox.askyesno("Confirm", f"Delete {count} record(s)?"):
            # Get row indices and sort descending
            indices = sorted([int(item) for item in selection], reverse=True)
            
            rows = self.controller.table_data.get("rows", [])
            for idx in indices:
                if idx < len(rows):
                    rows.pop(idx)
            
            self.controller.save_table_data()
            self._load_all_data()
            messagebox.showinfo("Success", f"✅ {count} record(s) deleted!")
    
    # ================================================================
    # EXPORT FUNCTIONS
    # ================================================================
    
    def _export_csv(self):
        """Export to CSV"""
        if not self.filtered_data:
            messagebox.showwarning("Warning", "No data available to export")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile=f"history_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if not filename:
            return
        
        try:
            import csv
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                # Write header
                writer.writerow(self.column_names)
                
                # Write data
                for record in self.filtered_data:
                    row = [record.get(col, "") for col in self.column_names]
                    writer.writerow(row)
            
            messagebox.showinfo("Success", f"✅ CSV Export successful!\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"CSV Export failed: {str(e)}")
    
    def _export_excel(self):
        """Export to Excel"""
        if not self.filtered_data:
            messagebox.showwarning("Warning", "No data available to export")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"history_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            import re
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "History Data"
            
            # Header styling
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, col_name in enumerate(self.column_names, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row_idx, record in enumerate(self.filtered_data, 2):
                for col_idx, col_name in enumerate(self.column_names, 1):
                    val = record.get(col_name, "")
                    # Convert to string and handle None
                    if val is None: val = ""
                    val_str = str(val)
                    # Clean control characters (Excel does not support them)
                    val_clean = re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', val_str)
                    ws.cell(row=row_idx, column=col_idx, value=val_clean)
            
            # Column width adjustment
            for col_idx, col_name in enumerate(self.column_names, 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 18
            
            wb.save(filename)
            messagebox.showinfo("Success", f"✅ Excel Export successful!\n{filename}")
            
        except ImportError:
            messagebox.showerror("Error", "Package 'openpyxl' not found.\nInstall: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def on_show(self):
        """Called when page is shown"""
        self._load_all_data()


# ================================================================
# EDIT ROW DIALOG
# ================================================================

class EditRowDialog(tk.Toplevel):
    """Dialog untuk edit row"""
    
    def __init__(self, parent, columns, values, colors):
        super().__init__(parent)
        
        self.colors = colors
        self.result = None
        self.columns = columns
        self.values = values
        self.entries = []
        
        self.title("✏️ Edit Data")
        self.geometry("500x600")
        self.configure(bg=colors["bg_card"])
        self.resizable(True, True)
        
        self._create_widgets()
        
        self.transient(parent)
        self.grab_set()
        
        # Center
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (500 // 2)
        y = (self.winfo_screenheight() // 2) - (600 // 2)
        self.geometry(f"+{x}+{y}")
    
    def _create_widgets(self):
        # Title
        tk.Label(
            self,
            text="✏️ Edit Data",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_primary"]
        ).pack(pady=15)
        
        # Scrollable content
        canvas = tk.Canvas(self, bg=self.colors["bg_card"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        content_frame = tk.Frame(canvas, bg=self.colors["bg_card"])
        
        content_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=content_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=20)
        scrollbar.pack(side="right", fill="y")
        
        # Fields
        for i, (col, val) in enumerate(zip(self.columns, self.values)):
            field_frame = tk.Frame(content_frame, bg=self.colors["bg_hover"], padx=10, pady=8)
            field_frame.pack(fill="x", pady=2)
            
            tk.Label(
                field_frame,
                text=col,
                font=("Segoe UI", 10, "bold"),
                bg=self.colors["bg_hover"],
                fg=self.colors["accent"],
                width=15,
                anchor="w"
            ).pack(side="left")
            
            entry = tk.Entry(
                field_frame,
                font=("Segoe UI", 10),
                bg=self.colors["bg_card"],
                fg=self.colors["text_primary"],
                insertbackground=self.colors["text_primary"],
                width=35
            )
            entry.insert(0, str(val))
            entry.pack(side="left", fill="x", expand=True, ipady=4)
            
            self.entries.append(entry)
        
        # Buttons
        btn_frame = tk.Frame(self, bg=self.colors["bg_card"])
        btn_frame.pack(pady=20)
        
        tk.Button(
            btn_frame,
            text="💾 Save",
            font=("Segoe UI", 11),
            bg=self.colors["accent_green"],
            fg="#1e1e2e",
            relief="flat",
            cursor="hand2",
            padx=25, pady=10,
            command=self._save
        ).pack(side="left", padx=10)
        
        tk.Button(
            btn_frame,
            text="❌ Cancel",
            font=("Segoe UI", 11),
            bg=self.colors["accent_red"],
            fg="#1e1e2e",
            relief="flat",
            cursor="hand2",
            padx=25, pady=10,
            command=self.destroy
        ).pack(side="left", padx=10)
    
    def _save(self):
        self.result = [entry.get() for entry in self.entries]
        self.destroy()