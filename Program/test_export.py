import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def export_excel_test():
    column_names = ["col1", "col2"]
    filtered_data = [{"col1": "A", "col2": "B"}]
    filename = "test_export.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "History Data"

    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(column_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(filtered_data, 2):
        for col_idx, col_name in enumerate(column_names, 1):
            val = record.get(col_name, "")
            if val is None: val = ""
            ws.cell(row=row_idx, column=col_idx, value=str(val))

    for col_idx, col_name in enumerate(column_names, 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 18

    wb.save(filename)
    print("Export successful!")

export_excel_test()
