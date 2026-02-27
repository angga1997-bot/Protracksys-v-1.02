"""
utils/excel_export.py - Excel Export with Images
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExporter:
    """Export data to Excel with images"""
    
    def __init__(self):
        self.has_openpyxl = HAS_OPENPYXL
    
    def export(self, data, columns, filename, include_images=True, image_column_indices=None):
        """
        Export data to Excel
        
        Args:
            data: List of dict or list
            columns: List of column names
            filename: Output filename
            include_images: Include images in export
            image_column_indices: List of column indices that contain images
        
        Returns:
            tuple: (success, message)
        """
        if not self.has_openpyxl:
            return self._export_csv(data, columns, filename)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Production Data"
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data
            row_height = 60 if include_images else 20
            image_column_indices = image_column_indices or []
            
            for row_idx, row_data in enumerate(data, 2):
                if isinstance(row_data, dict):
                    row_values = [row_data.get(col, "") for col in columns]
                else:
                    row_values = row_data
                
                ws.row_dimensions[row_idx].height = row_height
                
                for col_idx, value in enumerate(row_values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Check if this is an image column
                    if include_images and (col_idx - 1) in image_column_indices:
                        if value and os.path.exists(str(value)):
                            try:
                                img = XLImage(str(value))
                                img.width = 50
                                img.height = 50
                                ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")
                                cell.value = ""
                            except:
                                cell.value = os.path.basename(str(value))
                        else:
                            cell.value = str(value) if value else ""
                    else:
                        cell.value = str(value) if value else ""
            
            # Auto-fit column widths
            for col_idx, col_name in enumerate(columns, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(col_name) + 5)
            
            # Save
            wb.save(filename)
            
            return True, f"Export berhasil: {filename}"
            
        except Exception as e:
            return False, f"Export error: {str(e)}"
    
    def _export_csv(self, data, columns, filename):
        """Fallback to CSV export"""
        import csv
        
        try:
            csv_filename = filename.replace('.xlsx', '.csv')
            
            with open(csv_filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                
                for row_data in data:
                    if isinstance(row_data, dict):
                        row = [row_data.get(col, "") for col in columns]
                    else:
                        row = row_data
                    writer.writerow(row)
            
            return True, f"Export CSV berhasil: {csv_filename}\n(Install openpyxl untuk export Excel)"
            
        except Exception as e:
            return False, f"Export error: {str(e)}"